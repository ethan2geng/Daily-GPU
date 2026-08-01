#!/usr/bin/env python3
"""
Daily scraper for gpuperhour.com

Usage:
  python3 scrape.py --raw     # scrape only, save raw offers (runs at 10:00/10:15/10:30/10:45)
  python3 scrape.py           # scrape + compile best from all today's raws + charts + email (11:00)
"""

import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests

from charts import append_history, generate_charts_excel, send_email

SCRIPT_DIR = Path(__file__).parent
SAVE_DIR = SCRIPT_DIR / "output"
SAVE_DIR.mkdir(parents=True, exist_ok=True)

RAW_DIR    = SCRIPT_DIR / "raw"
RAW_DIR.mkdir(exist_ok=True)


# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------

# gpuperhour.com's frontend reads from this public JSON API — hitting it
# directly avoids driving a browser through gpuperhour.com's own pagination,
# which has a client-side hydration bug that silently drops every page past
# the first (the "Next" click fires the right request but the table never
# re-renders).
API_URL = "https://api.gpuindexes.com/api/offers"
API_PARAMS = {
    "sortBy": "priceHourly",
    "sortOrder": "asc",
    "available": "true",
    "liveStockOnly": "true",
    "pricingType": "on-demand",
    "securityTier": "secure",
    "limit": 100,  # API max
}


def fetch_page(page_num: int, retries: int = 3) -> dict:
    params = {**API_PARAMS, "page": page_num}
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(API_URL, params=params, timeout=20)
            resp.raise_for_status()
            return resp.json()
        except requests.RequestException as e:
            if attempt == retries:
                raise
            print(f"  Page {page_num}: request failed ({e}), retrying...", flush=True)
            time.sleep(2 * attempt)


def scrape_all() -> list[dict]:
    all_offers: list[dict] = []
    page_num = 1
    while True:
        payload = fetch_page(page_num)
        for o in payload["data"]:
            provider  = o.get("provider")
            gpu_model = o.get("gpu", {}).get("name")
            price     = o.get("pricePerGpu")
            if provider and gpu_model and price is not None:
                all_offers.append({"provider": provider, "gpu_model": gpu_model, "price": price})

        print(f"  Page {page_num:2d}: {len(payload['data']):3d} rows  (total: {len(all_offers)})", flush=True)

        if not payload.get("pagination", {}).get("hasMore"):
            break
        page_num += 1

    return all_offers


# ---------------------------------------------------------------------------
# Raw file helpers
# ---------------------------------------------------------------------------

def save_raw(offers: list[dict]) -> Path:
    """Save raw offers to a timestamped CSV in raw/."""
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = RAW_DIR / f"{timestamp}.csv"
    pd.DataFrame(offers).to_csv(path, index=False)
    print(f"Raw saved → {path}  ({len(offers)} offers)", flush=True)
    return path


def compile_best_offers() -> list[dict]:
    """
    Combine today's raw scrapes into one best-effort offer list.
    For each (gpu_model, provider) pair, take the price from the most
    recent scrape that has data (11:00 → 10:45 → 10:30 → 10:15 → 10:00).
    """
    today = datetime.now().strftime("%Y-%m-%d")
    files = sorted(RAW_DIR.glob(f"{today}_*.csv"), reverse=True)  # most recent first

    if not files:
        print("No raw files found for today.", flush=True)
        return []

    print(f"Compiling from {len(files)} raw scrape(s): {[f.name for f in files]}", flush=True)

    seen: dict[tuple, float] = {}   # (gpu_model, provider) → price
    for f in files:
        df = pd.read_csv(f)
        for _, row in df.iterrows():
            key = (row["gpu_model"], row["provider"])
            if key not in seen:
                seen[key] = row["price"]

    offers = [{"gpu_model": k[0], "provider": k[1], "price": v} for k, v in seen.items()]
    print(f"Compiled {len(offers)} unique (model, provider) offers.", flush=True)
    return offers


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def save_excel(offers: list[dict]) -> Path:
    if not offers:
        raise RuntimeError("No offers collected")

    df = pd.DataFrame(offers)
    print(f"Rows: {len(df)}  |  providers: {df['provider'].nunique()}  |  GPUs: {df['gpu_model'].nunique()}", flush=True)

    pivot = (
        df.groupby(["gpu_model", "provider"])["price"]
        .min()
        .unstack("provider")
    )
    pivot.index.name = "GPU Model"
    pivot = pivot.sort_index()

    today = datetime.now().strftime("%Y-%m-%d")
    path = SAVE_DIR / f"gpu_prices_{today}.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Prices")
        ws = writer.sheets["Prices"]

        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        index_font  = Font(bold=True, size=10)

        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for i, cell in enumerate(row):
                if i == 0:
                    cell.font = index_font
                elif cell.value is not None:
                    cell.number_format = '"$"#,##0.00'

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 28)

        ws.freeze_panes = "B2"

    print(f"Saved → {path}", flush=True)
    return path


# ---------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------

def main():
    raw_only = "--raw" in sys.argv
    now = datetime.now().strftime("%Y-%m-%d %H:%M %Z")

    if raw_only:
        print(f"=== Raw Scrape  {now} ===", flush=True)
        offers = scrape_all()
        save_raw(offers)
        print(f"=== Raw done ===", flush=True)
    else:
        print(f"=== Full Run  {now} ===", flush=True)
        offers = scrape_all()
        save_raw(offers)                          # save 11:00 raw too
        best = compile_best_offers()              # merge all 5 scrapes
        daily_path  = save_excel(best)
        hist        = append_history(best)
        charts_path = generate_charts_excel(hist)
        send_email(charts_path, daily_path, history_df=hist)
        print(f"=== Done: {daily_path} ===", flush=True)


if __name__ == "__main__":
    main()
